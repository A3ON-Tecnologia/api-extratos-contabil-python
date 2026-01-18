"""
Serviço de auditoria e logging em planilha Excel.

Responsável por registrar o resultado de cada processamento
na planilha de LOG.
"""

import logging
from datetime import datetime
from pathlib import Path
from threading import Lock
import time

from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.config import get_settings
from app.schemas.api import ProcessingStatus

logger = logging.getLogger(__name__)


class AuditService:
    """Serviço de auditoria via planilha Excel."""
    
    # Lock para escrita concorrente
    _write_lock = Lock()
    
    # Colunas esperadas no LOG
    COLUMNS = [
        "DATA",
        "NOME DO CLIENTE",
        "TIPO EXTRATO",
        "ANO",
        "MÊS",
        "STATUS",
        "NOME ARQUIVO FINAL",
    ]
    
    def __init__(self):
        """Inicializa o serviço."""
        self.settings = get_settings()
    
    def log_result(
        self,
        nome_cliente: str | None,
        tipo_extrato: str | None,
        ano: int | None,
        mes: int | None,
        status: ProcessingStatus,
        nome_arquivo_final: str | None,
    ) -> bool:
        """
        Registra o resultado do processamento na planilha de LOG.
        
        Adiciona uma nova linha (append) sem sobrescrever dados existentes.
        
        Args:
            nome_cliente: Nome do cliente identificado ou "NÃO IDENTIFICADO"
            tipo_extrato: Tipo do documento (retornado pela LLM)
            ano: Ano do documento
            mes: Mês do documento
            status: Status do processamento (SUCESSO, NÃO IDENTIFICADO, FALHA)
            nome_arquivo_final: Caminho/nome do arquivo salvo
            
        Returns:
            True se conseguiu registrar, False caso contrário
        """
        with self._write_lock:
            return self._write_log_entry(
                nome_cliente=nome_cliente or "NÃO IDENTIFICADO",
                tipo_extrato=tipo_extrato or "Não identificado",
                ano=ano,
                mes=mes,
                status=status.value,
                nome_arquivo_final=nome_arquivo_final or "",
            )
    
    def _write_log_entry(
        self,
        nome_cliente: str,
        tipo_extrato: str,
        ano: int | None,
        mes: int | None,
        status: str,
        nome_arquivo_final: str,
    ) -> bool:
        """
        Escreve uma entrada no arquivo de LOG.
        
        Tenta múltiplas vezes em caso de arquivo bloqueado.
        """
        log_path = self.settings.log_excel_path
        max_retries = 3
        retry_delay = 1  # segundos
        
        for attempt in range(max_retries):
            try:
                # Tenta abrir ou criar a planilha
                workbook, worksheet = self._open_or_create_log(log_path)
                
                # Dados da nova linha
                now = datetime.now()
                row_data = [
                    now.strftime("%d/%m/%Y %H:%M:%S"),  # DATA
                    nome_cliente,                       # NOME DO CLIENTE
                    tipo_extrato,                       # TIPO EXTRATO
                    str(ano) if ano else "",            # ANO
                    str(mes) if mes else "",            # MÊS
                    status,                             # STATUS
                    nome_arquivo_final,                 # NOME ARQUIVO FINAL
                ]
                
                # Adiciona a linha
                worksheet.append(row_data)
                
                # Salva o arquivo
                workbook.save(log_path)
                workbook.close()
                
                logger.info(
                    f"Log registrado: {nome_cliente} - {status}"
                )
                return True
                
            except PermissionError:
                # Arquivo provavelmente aberto por outro processo
                logger.warning(
                    f"Arquivo de LOG bloqueado, tentativa {attempt + 1}/{max_retries}"
                )
                
                if attempt < max_retries - 1:
                    time.sleep(retry_delay)
                    continue
                
                # Última tentativa - salvar em arquivo alternativo
                return self._write_to_fallback(
                    log_path,
                    nome_cliente,
                    tipo_extrato,
                    ano,
                    mes,
                    status,
                    nome_arquivo_final
                )
                
            except Exception as e:
                logger.error(f"Erro ao escrever no LOG: {e}")
                return False
        
        return False
    
    def _open_or_create_log(self, log_path: Path) -> tuple:
        """
        Abre a planilha de LOG existente ou cria uma nova.
        
        Returns:
            Tuple (workbook, worksheet)
        """
        if log_path.exists():
            try:
                workbook = load_workbook(log_path)
                worksheet = workbook.active
                return workbook, worksheet
            except InvalidFileException:
                logger.warning("Arquivo de LOG corrompido, criando novo")
        
        # Cria nova planilha
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "LOG"
        
        # Adiciona cabeçalhos
        worksheet.append(self.COLUMNS)
        
        # Garante que o diretório existe
        log_path.parent.mkdir(parents=True, exist_ok=True)
        
        return workbook, worksheet
    
    def _write_to_fallback(
        self,
        original_path: Path,
        nome_cliente: str,
        tipo_extrato: str,
        ano: int | None,
        mes: int | None,
        status: str,
        nome_arquivo_final: str,
    ) -> bool:
        """
        Escreve em arquivo alternativo quando o principal está bloqueado.
        
        Cria um arquivo com timestamp no nome.
        """
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            fallback_name = f"LOG_FALLBACK_{timestamp}.xlsx"
            fallback_path = original_path.parent / fallback_name
            
            logger.warning(f"Usando arquivo de fallback: {fallback_path}")
            
            # Cria nova planilha
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "LOG"
            
            # Adiciona cabeçalhos
            worksheet.append(self.COLUMNS)
            
            # Adiciona a linha
            now = datetime.now()
            row_data = [
                now.strftime("%d/%m/%Y %H:%M:%S"),
                nome_cliente,
                tipo_extrato,
                str(ano) if ano else "",
                str(mes) if mes else "",
                status,
                nome_arquivo_final,
            ]
            worksheet.append(row_data)
            
            # Salva
            workbook.save(fallback_path)
            workbook.close()
            
            logger.info(f"Log de fallback salvo: {fallback_path}")
            return True
            
        except Exception as e:
            logger.error(f"Erro ao escrever log de fallback: {e}")
            return False
    
    def merge_fallback_logs(self) -> int:
        """
        Mescla logs de fallback no arquivo principal.
        
        Deve ser chamado manualmente ou em um processo de manutenção.
        
        Returns:
            Número de entradas mescladas
        """
        log_dir = self.settings.log_excel_path.parent
        fallback_files = list(log_dir.glob("LOG_FALLBACK_*.xlsx"))
        
        if not fallback_files:
            return 0
        
        merged_count = 0
        
        with self._write_lock:
            try:
                # Abre o arquivo principal
                main_workbook, main_worksheet = self._open_or_create_log(
                    self.settings.log_excel_path
                )
                
                # Processa cada arquivo de fallback
                for fallback_file in fallback_files:
                    try:
                        fb_workbook = load_workbook(fallback_file)
                        fb_worksheet = fb_workbook.active
                        
                        # Pula a primeira linha (cabeçalho)
                        for row in fb_worksheet.iter_rows(
                            min_row=2, 
                            values_only=True
                        ):
                            if any(row):  # Ignora linhas vazias
                                main_worksheet.append(list(row))
                                merged_count += 1
                        
                        fb_workbook.close()
                        
                        # Remove o arquivo de fallback
                        fallback_file.unlink()
                        logger.info(f"Fallback mesclado e removido: {fallback_file}")
                        
                    except Exception as e:
                        logger.error(f"Erro ao mesclar {fallback_file}: {e}")
                
                # Salva o arquivo principal
                main_workbook.save(self.settings.log_excel_path)
                main_workbook.close()
                
            except Exception as e:
                logger.error(f"Erro ao mesclar fallbacks: {e}")
        
        return merged_count

    def get_recent_logs(self, limit: int = 100) -> list[dict]:
        """
        Lê as últimas entradas do arquivo de log.
        
        Returns:
            Lista de dicionários com os logs recentes.
        """
        if not self.settings.log_excel_path.exists():
            return []

        try:
            # Abre como read-only para velocidade
            workbook = load_workbook(self.settings.log_excel_path, read_only=True, data_only=True)
            worksheet = workbook.active
            
            # iter_rows em read_only não permite indexação reversa fácil sem ler tudo
            # Mas vamos ler tudo pois é texto simples e não deve ser gigante
            all_rows = list(worksheet.iter_rows(min_row=2, values_only=True))
            
            # Pega os ultimos 'limit'
            recent_rows = all_rows[-limit:]
            
            # Inverte para mostrar o mais recente primeiro (topo da lista)
            recent_rows.reverse()
            
            result = []
            for row in recent_rows:
                if not row or not any(row): continue
                
                # Mapeia colunas: 0=DATA, 1=CLIENTE, 2=TIPO, 3=ANO, 4=MES, 5=STATUS, 6=PATH
                # Garante que temos colunas suficientes
                if len(row) < 7: continue

                full_path = str(row[6]) if row[6] else ""
                filename = Path(full_path).name if full_path else "Desconhecido"
                
                result.append({
                    "timestamp": datetime.now().isoformat(), # Placeholder, data real está formatada string
                    "data_hora_formatada": str(row[0]),
                    "cliente": str(row[1]),
                    "tipo": str(row[2]),
                    "status": str(row[5]),
                    "filename": filename,
                    "full_path": full_path,
                    "periodo": f"{row[4]}/{row[3]}" if row[3] and row[4] else ""
                })
            
            workbook.close()
            return result
            
        except Exception as e:
            logger.error(f"Erro ao ler histórico de logs: {e}")
            return []
