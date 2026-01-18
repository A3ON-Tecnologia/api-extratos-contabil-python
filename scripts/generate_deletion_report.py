
import logging
from pathlib import Path
from openpyxl import load_workbook
import datetime

# Setup
logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)

def generate_report():
    log_path = Path(r"J:\JP Digital\000 - AUTOMAÇÕES\LOGS-SUCESSO-FALHA.xlsx")
    output_file = Path("arquivos_e_pastas_excluidos.txt")
    target_date_str = "13/01/2026"
    
    if not log_path.exists():
        print(f"Erro: Arquivo de log não encontrado em {log_path}")
        return

    print(f"Gerando relatório de exclusão baseado no log de {target_date_str}...")
    
    try:
        wb = load_workbook(log_path, read_only=True)
        ws = wb.active
        
        processed_files = []
        
        # 1. Coleta arquivos do dia 13/01
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            
            cell_date = row[0]
            if isinstance(cell_date, datetime.datetime):
                date_str = cell_date.strftime("%d/%m/%Y")
            else:
                date_str = str(cell_date).split()[0]
            
            if target_date_str in date_str:
                file_path_str = row[6]
                if file_path_str and str(file_path_str).strip():
                    processed_files.append(Path(file_path_str))
        
        if not processed_files:
            print("Nenhum registro encontrado para esta data.")
            return

        # 2. Analisa Pastas Excluídas (aquelas que não existem mais)
        # Agrupamos por pasta para não repetir
        unique_dirs = set(f.parent for f in processed_files)
        deleted_dirs = []
        kept_dirs = []
        
        for folder in unique_dirs:
            if not folder.exists():
                deleted_dirs.append(folder)
            else:
                kept_dirs.append(folder)
        
        # 3. Escreve o Relatório TXT
        with open(output_file, "w", encoding="utf-8") as f:
            f.write(f"RELATÓRIO DE EXCLUSÃO - DATA: {target_date_str}\n")
            f.write("==================================================\n\n")
            
            f.write(f"RESUMO:\n")
            f.write(f"- Arquivos processados/excluídos: {len(processed_files)}\n")
            f.write(f"- Pastas completas excluídas: {len(deleted_dirs)}\n")
            f.write(f"- Pastas mantidas: {len(kept_dirs)}\n\n")
            
            f.write("--------------------------------------------------\n")
            f.write("PASTAS EXCLUÍDAS (21 identificadas anteriormente):\n")
            f.write("--------------------------------------------------\n")
            if deleted_dirs:
                for d in sorted(deleted_dirs):
                    f.write(f"[X] {d}\n")
            else:
                f.write("(Nenhuma pasta encontrada como excluída ou todas já existiam antes)\n")
            
            f.write("\n")
            f.write("--------------------------------------------------\n")
            f.write("ARQUIVOS EXCLUÍDOS:\n")
            f.write("--------------------------------------------------\n")
            for file_path in sorted(processed_files):
                f.write(f"[X] {file_path.name}\n")
                f.write(f"    Caminho original: {file_path}\n")
            
            f.write("\n")
            f.write("--------------------------------------------------\n")
            f.write("FIM DO RELATÓRIO\n")
            
        print(f"Relatório gerado com sucesso: {output_file.absolute()}")

    except Exception as e:
        logger.error(f"Erro ao gerar relatório: {e}")

if __name__ == "__main__":
    generate_report()
