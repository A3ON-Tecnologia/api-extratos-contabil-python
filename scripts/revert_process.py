
import logging
from pathlib import Path
from openpyxl import load_workbook
import datetime
import sys

# Configuração de Logs
logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)

def revert_process():
    log_path = Path(r"J:\JP Digital\000 - AUTOMAÇÕES\LOGS-SUCESSO-FALHA.xlsx")
    target_date_str = "13/01/2026"
    
    print(f"--- INICIANDO PROCESSO DE REVERSÃO ---")
    print(f"Data alvo: {target_date_str}")
    
    if not log_path.exists():
        logger.error(f"Arquivo de log não encontrado: {log_path}")
        return

    try:
        wb = load_workbook(log_path, read_only=True)
        ws = wb.active
        
        files_to_remove = []
        
        # 1. Coleta arquivos do log
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            
            cell_date = row[0]
            if isinstance(cell_date, datetime.datetime):
                date_str = cell_date.strftime("%d/%m/%Y")
            else:
                date_str = str(cell_date).split()[0]
            
            if target_date_str in date_str:
                file_path_str = row[6]
                if file_path_str and str(file_path_str).strip():
                    files_to_remove.append(Path(file_path_str))
        
        if not files_to_remove:
            print("Nenhum arquivo encontrado para remoção.")
            return

        print(f"Arquivos identificados: {len(files_to_remove)}")
        
        # 2. Remove arquivos
        removed_files_count = 0
        parents_to_check = set()
        
        for file_path in files_to_remove:
            if file_path.exists():
                try:
                    parents_to_check.add(file_path.parent)
                    file_path.unlink()
                    print(f"Removido: {file_path.name}")
                    removed_files_count += 1
                except Exception as e:
                    print(f"Erro ao remover {file_path.name}: {e}")
            else:
                # Se não existe, ainda adicionamos o pai para verificação de pasta vazia
                # caso o usuário tenha deletado o arquivo manualmente mas deixado a pasta
                parents_to_check.add(file_path.parent)
                print(f"Ignorado (não existe): {file_path.name}")

        print(f"\nArquivos removidos com sucesso: {removed_files_count}")

        # 3. Remove pastas vazias
        print(f"\nVerificando {len(parents_to_check)} pastas para limpeza...")
        removed_dirs_count = 0
        
        for folder in parents_to_check:
            if not folder.exists():
                continue
                
            # Verifica se está vazia
            if not any(folder.iterdir()):
                try:
                    folder.rmdir()
                    print(f"Pasta vazia removida: {folder}")
                    removed_dirs_count += 1
                except Exception as e:
                    print(f"Erro ao remover pasta {folder}: {e}")
            else:
                print(f"Pasta mantida (não está vazia): {folder}")
        
        print(f"\n--- REVERSÃO CONCLUÍDA ---")
        print(f"Total de arquivos apagados: {removed_files_count}")
        print(f"Total de pastas apagadas: {removed_dirs_count}")

    except Exception as e:
        logger.error(f"Erro fatal durante a reversão: {e}")

if __name__ == "__main__":
    revert_process()
