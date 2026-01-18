"""
Script para gerar relatório Excel de reversão de processamentos.
Cria um XLSX com todas as informações antes de reverter.
"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.database import SessionLocal
from app.models.extrato_log import ExtratoLog
from sqlalchemy import desc


def gerar_relatorio_reversao(quantidade=68, output_path=None):
    """
    Gera um arquivo Excel com os processamentos que serão revertidos.
    
    Args:
        quantidade: Número de processamentos mais recentes para incluir
        output_path: Caminho do arquivo de saída (opcional)
    
    Returns:
        Caminho do arquivo gerado
    """
    db = SessionLocal()
    
    # Busca os últimos N processamentos
    logs = db.query(ExtratoLog).order_by(desc(ExtratoLog.id)).limit(quantidade).all()
    
    # Cria o workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Reversao"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Cabeçalhos
    headers = [
        "ID",
        "DATA/HORA",
        "CLIENTE",
        "COD CLIENTE",
        "CNPJ",
        "BANCO",
        "TIPO DOCUMENTO",
        "AGENCIA",
        "CONTA",
        "ANO",
        "MES",
        "STATUS",
        "METODO",
        "CONFIANCA IA",
        "ARQUIVO ORIGINAL",
        "ARQUIVO SALVO",
        "EXISTE NO DISCO",
        "ERRO",
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Dados
    arquivos_existentes = 0
    arquivos_nao_existentes = 0
    
    for row, log in enumerate(logs, 2):
        # Verifica se arquivo existe
        arquivo_existe = False
        if log.arquivo_salvo:
            path = Path(log.arquivo_salvo)
            arquivo_existe = path.exists()
            if arquivo_existe:
                arquivos_existentes += 1
            else:
                arquivos_nao_existentes += 1
        
        dados = [
            log.id,
            log.processado_em.strftime("%d/%m/%Y %H:%M:%S") if log.processado_em else "",
            log.cliente_nome or "",
            log.cliente_cod or "",
            log.cliente_cnpj or "",
            log.banco or "",
            log.tipo_documento or "",
            log.agencia or "",
            log.conta or "",
            log.ano or "",
            log.mes or "",
            log.status or "",
            log.metodo_identificacao or "",
            f"{log.confianca_ia}%" if log.confianca_ia else "",
            log.arquivo_original or "",
            log.arquivo_salvo or "",
            "SIM" if arquivo_existe else "NAO",
            log.erro or "",
        ]
        
        for col, valor in enumerate(dados, 1):
            cell = ws.cell(row=row, column=col, value=valor)
            cell.border = border
            
            # Aplica cor se arquivo existe
            if col == 17:  # Coluna "EXISTE NO DISCO"
                if valor == "SIM":
                    cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                else:
                    cell.fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
            
            # Cor do status
            if col == 12:  # Coluna STATUS
                if valor == "SUCESSO":
                    cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                elif valor == "NAO_IDENTIFICADO":
                    cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                elif valor == "FALHA":
                    cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    
    # Ajusta largura das colunas
    column_widths = [6, 18, 40, 12, 18, 15, 18, 10, 15, 6, 6, 18, 18, 12, 30, 60, 15, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Adiciona resumo em uma nova aba
    ws_resumo = wb.create_sheet("Resumo")
    
    resumo_data = [
        ["RELATORIO DE REVERSAO", ""],
        ["", ""],
        ["Data/Hora do Relatório", datetime.now().strftime("%d/%m/%Y %H:%M:%S")],
        ["", ""],
        ["Total de Registros", len(logs)],
        ["Arquivos que EXISTEM no disco", arquivos_existentes],
        ["Arquivos que NAO existem no disco", arquivos_nao_existentes],
        ["", ""],
        ["STATUS", "QUANTIDADE"],
    ]
    
    # Conta por status
    status_count = {}
    for log in logs:
        status = log.status or "DESCONHECIDO"
        status_count[status] = status_count.get(status, 0) + 1
    
    for status, count in status_count.items():
        resumo_data.append([status, count])
    
    for row, dados in enumerate(resumo_data, 1):
        for col, valor in enumerate(dados, 1):
            cell = ws_resumo.cell(row=row, column=col, value=valor)
            if row == 1:
                cell.font = Font(bold=True, size=14)
    
    ws_resumo.column_dimensions['A'].width = 35
    ws_resumo.column_dimensions['B'].width = 20
    
    # Define nome do arquivo
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = Path(__file__).parent.parent / f"REVERSAO_{timestamp}.xlsx"
    
    # Salva
    wb.save(output_path)
    
    db.close()
    
    print(f"=" * 60)
    print(f"RELATORIO DE REVERSAO GERADO")
    print(f"=" * 60)
    print(f"Arquivo: {output_path}")
    print(f"Total de registros: {len(logs)}")
    print(f"Arquivos existentes no disco: {arquivos_existentes}")
    print(f"Arquivos que NAO existem: {arquivos_nao_existentes}")
    print(f"=" * 60)
    
    return output_path


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Gera relatorio Excel de reversao")
    parser.add_argument("--quantidade", "-n", type=int, default=68, help="Quantidade de processamentos")
    parser.add_argument("--output", "-o", type=str, help="Caminho do arquivo de saida")
    
    args = parser.parse_args()
    
    output = args.output
    if output:
        output = Path(output)
    
    gerar_relatorio_reversao(args.quantidade, output)
