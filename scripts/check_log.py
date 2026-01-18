
import logging
from pathlib import Path
from openpyxl import load_workbook
import datetime

# Configuração de Logs
logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)

import sys

def check_impact():
    log_path = Path(r"J:\JP Digital\000 - AUTOMAÇÕES\LOGS-SUCESSO-FALHA.xlsx")
    target_date_str = "13/01/2026"
    
    # Redireciona stdout para arquivo
    with open("reversion_report.txt", "w", encoding="utf-8") as f:
        original_stdout = sys.stdout
        sys.stdout = f
        
        try:
            if not log_path.exists():
                logger.error(f"Arquivo de log não encontrado em: {log_path}")
                return

            print(f"--- RELATÓRIO DE SIMULAÇÃO DE REVERSÃO ---")
            print(f"Lendo log: {log_path}")
            print(f"Data alvo: {target_date_str}\n")
            
            wb = load_workbook(log_path, read_only=True)
            ws = wb.active
            
            files_to_remove = []
            
            # Itera sobre as linhas do Excel
            # Colunas (index): 0=DATA, 6=NOME ARQUIVO FINAL
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                    
                cell_date = row[0]
                # Normaliza a data para string dd/mm/yyyy
                if isinstance(cell_date, datetime.datetime):
                    date_str = cell_date.strftime("%d/%m/%Y")
                else:
                    date_str = str(cell_date).split()[0]
                
                # Checa se é a data desejada e se tem arquivo salvo
                if target_date_str in date_str:
                    file_path_str = row[6]
                    if file_path_str and str(file_path_str).strip():
                        files_to_remove.append(Path(file_path_str))
            
            if not files_to_remove:
                print(f"Nenhum arquivo encontrado no log para a data {target_date_str}.")
                return

            # Agrupa arquivos por pasta pai para análise
            files_by_dir = {}
            for f in files_to_remove:
                parent = f.parent
                if parent not in files_by_dir:
                    files_by_dir[parent] = []
                files_by_dir[parent].append(f)
                
            total_files = len(files_to_remove)
            print(f"Total de arquivos identificados no log para remoção: {total_files}\n")
            
            print(f"--- ANÁLISE DETALHADA POR PASTA ---")
            
            dirs_to_remove = []
            dirs_to_keep = []
            
            for directory, files in files_by_dir.items():
                if not directory.exists():
                    print(f"[!] Pasta não existe mais: {directory} ({len(files)} arquivos listados no log)")
                    continue
                    
                # Lista todos os arquivos que realmente existem na pasta hoje
                try:
                    all_files_in_dir = list(directory.glob('*'))
                    existing_target_files = [f for f in files if f.exists()]
                    
                    # Simula a remoção
                    remaining_count = len(all_files_in_dir) - len(existing_target_files)
                    
                    print(f"Pasta: {directory}")
                    print(f"  - Arquivos gerados neste dia: {len(files)} ({len(existing_target_files)} existem no disco)")
                    
                    if remaining_count == 0:
                        status = " [SERÁ REMOVIDA COMPLETAMENTE]"
                        dirs_to_remove.append(directory)
                    else:
                        status = f" [SERÁ MANTIDA - CONTÉM OUTROS {remaining_count} ARQUIVOS]"
                        dirs_to_keep.append(directory)
                    
                    print(f"  - Status pós-reversão: {status}")
                    # Lista amostra de arquivos
                    for f in files[:3]:
                        print(f"    Ex: {f.name}")
                    if len(files) > 3:
                        print(f"    ... e mais {len(files)-3} arquivos")
                    print("")

                except Exception as e:
                    print(f"Erro ao acessar pasta {directory}: {e}")

            print(f"--- RESUMO FINAL ---")
            print(f"Arquivos para deletar: {total_files}")
            print(f"Pastas que ficarão vazias e serão excluídas: {len(dirs_to_remove)}")
            print(f"Pastas que serão mantidas (conteúdo misto): {len(dirs_to_keep)}")
            print("\nNENHUMA AÇÃO FOI TOMADA. ISSO É APENAS UM RELATÓRIO.")

        except Exception as e:
            logger.error(f"Erro ao processar arquivo de log: {e}")
        finally:
            sys.stdout = original_stdout
            print("Relatório salvo em reversion_report.txt")

if __name__ == "__main__":
    check_impact()
